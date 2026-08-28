import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SRC = "/tmp/claude-0/-home-user-saas-landing/cbf105b2-215c-546e-a953-9c154a810568/scratchpad/bank.json"
OUT = "/home/user/saas-landing/toefl-platform/docs/TOEFL-Reading-Question-Bank.xlsx"

d = json.load(open(SRC))

NAVY   = "0B1B2B"
HEADBG = "17395A"
AMBER  = "E8A33D"
BAND   = "EDF2F6"
GREEN  = "1C7A5B"

BASE   = Font(name="Arial", size=10, color="14202B")
HEAD   = Font(name="Arial", size=10, bold=True, color="FFFFFF")
TITLE  = Font(name="Arial", size=16, bold=True, color="0B1B2B")
SUB    = Font(name="Arial", size=10, color="47596A")
MONO   = Font(name="Consolas", size=10, color="14202B")
KEY    = Font(name="Arial", size=10, bold=True, color=GREEN)
BOLD   = Font(name="Arial", size=10, bold=True, color="14202B")

head_fill = PatternFill("solid", fgColor=HEADBG)
band_fill = PatternFill("solid", fgColor=BAND)
thin = Side(style="thin", color="C9D6E0")
box = Border(left=thin, right=thin, top=thin, bottom=thin)

top_wrap = Alignment(vertical="top", wrap_text=True)
top_left = Alignment(vertical="top", horizontal="left")
centre = Alignment(vertical="center", horizontal="center")

wb = Workbook()

def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = box
    ws.row_dimensions[row].height = 28
    ws.freeze_panes = ws.cell(row=row + 1, column=1)

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w

# ── Overview ───────────────────────────────────────────────────────────────
ov = wb.active
ov.title = "Overview"
ov.sheet_view.showGridLines = False
widths(ov, {"A": 3, "B": 34, "C": 14, "D": 14, "E": 60})

ov["B2"] = "TOEFL® Reading — Practice Question Bank"
ov["B2"].font = TITLE
ov["B3"] = "159 questions across the three passage types of the TOEFL iBT Reading section."
ov["B3"].font = SUB

hdr = 5
for i, h in enumerate(["Section", "Sets", "Questions", "What it tests"], start=2):
    ov.cell(row=hdr, column=i, value=h)
style_header(ov, hdr, 5)

# Bounded ranges, no INDIRECT: the counts stay live but cost nothing to recompute.
n_cloze = sum(len(s["blanks"]) for s in d["cloze"])
n_daily = sum(len(s["questions"]) for s in d["reading"] if s["section"] == "daily-life")
n_acad = sum(len(s["questions"]) for s in d["reading"] if s["section"] == "academic")

rows = [
    ("Complete the Words", n_cloze, "Ten short paragraphs with letters removed; type the missing letters."),
    ("Read in Daily Life", n_daily, "Emails, announcements, and text chains. Purpose, detail, and inference."),
    ("Read an Academic Passage", n_acad, "Three-paragraph passages: vocabulary, purpose, insert text, sentence select."),
]

r = hdr + 1
for name, count, blurb in rows:
    last = count + 1
    ov.cell(row=r, column=2, value=name).font = BOLD
    # Counts read the data sheets, so this summary cannot drift from the content.
    ov.cell(row=r, column=3,
            value="=SUMPRODUCT(1/COUNTIF('%s'!B2:B%d,'%s'!B2:B%d))" % (name, last, name, last))
    ov.cell(row=r, column=4, value="=COUNT('%s'!A2:A%d)" % (name, last))
    ov.cell(row=r, column=5, value=blurb)
    for c in range(2, 6):
        cell = ov.cell(row=r, column=c)
        cell.border = box
        cell.alignment = top_wrap if c == 5 else (centre if c in (3, 4) else top_left)
        if cell.font is None or c not in (2,):
            cell.font = BASE
    ov.row_dimensions[r].height = 30
    r += 1

ov.cell(row=r, column=2, value="Total").font = BOLD
ov.cell(row=r, column=3, value=f"=SUM(C{hdr+1}:C{r-1})").font = BOLD
ov.cell(row=r, column=4, value=f"=SUM(D{hdr+1}:D{r-1})").font = BOLD
for c in range(2, 6):
    ov.cell(row=r, column=c).border = box
    ov.cell(row=r, column=c).fill = PatternFill("solid", fgColor="FDF0D8")
    if c in (3, 4):
        ov.cell(row=r, column=c).alignment = centre

# Full tests
r += 3
ov.cell(row=r, column=2, value="Full tests").font = Font(name="Arial", size=12, bold=True, color="0B1B2B")
ov.cell(row=r + 1, column=2, value="Each pairs one passage of every type under a single clock — the shape of the real section.").font = SUB
r += 3
for i, h in enumerate(["Test", "Minutes", "Questions", "Passages"], start=2):
    ov.cell(row=r, column=i, value=h)
style_header(ov, r, 5)
ov.freeze_panes = "A6"
r += 1
for e in d["exams"]:
    ov.cell(row=r, column=2, value=e["title"]).font = BOLD
    ov.cell(row=r, column=3, value=e["minutes"]).font = BASE
    n = 0
    for sid in e["setIds"]:
        if sid.startswith("cw-"):
            n += 10
        else:
            n += next(len(s["questions"]) for s in d["reading"] if s["id"] == sid)
    ov.cell(row=r, column=4, value=n).font = BASE
    ov.cell(row=r, column=5, value=e["blurb"]).font = BASE
    for c in range(2, 6):
        ov.cell(row=r, column=c).border = box
        ov.cell(row=r, column=c).alignment = centre if c in (3, 4) else (top_wrap if c == 5 else top_left)
    r += 1

r += 2
note = ov.cell(row=r, column=2, value=(
    "Source: 100 Practice Questions for the TOEFL® Reading Section (TST Prep, 2026 edition). "
    "Question numbers match that book. TOEFL® is a registered trademark of Educational Testing "
    "Service (ETS); this material is not endorsed or approved by ETS."))
note.font = SUB
note.alignment = top_wrap
ov.merge_cells(start_row=r, start_column=2, end_row=r + 2, end_column=5)

# ── Complete the Words ─────────────────────────────────────────────────────
cw = wb.create_sheet("Complete the Words")
cw.sheet_view.showGridLines = False
cols = ["#", "Set", "Paragraph", "Stem shown", "Missing letters", "Complete word"]
for i, h in enumerate(cols, start=1):
    cw.cell(row=1, column=i, value=h)
style_header(cw, 1, len(cols))
widths(cw, {"A": 6, "B": 8, "C": 26, "D": 14, "E": 18, "F": 20})

r = 2
for s in d["cloze"]:
    for b in s["blanks"]:
        cw.cell(row=r, column=1, value=b["number"]).font = BASE
        cw.cell(row=r, column=2, value=s["index"]).font = BASE
        cw.cell(row=r, column=3, value=s["title"]).font = BASE
        cw.cell(row=r, column=4, value=b["stem"]).font = MONO
        cw.cell(row=r, column=5, value=b["answer"]).font = KEY
        # Derived, not duplicated: the word is the stem plus the letters.
        cw.cell(row=r, column=6, value=f"=D{r}&E{r}").font = MONO
        for c in range(1, len(cols) + 1):
            cell = cw.cell(row=r, column=c)
            cell.border = box
            cell.alignment = centre if c in (1, 2) else top_left
            if s["index"] % 2 == 0:
                cell.fill = band_fill
        r += 1
cw.auto_filter.ref = f"A1:F{r-1}"

# ── Cloze paragraphs ───────────────────────────────────────────────────────
cp = wb.create_sheet("Cloze Paragraphs")
cp.sheet_view.showGridLines = False
for i, h in enumerate(["Set", "Title", "Blanks", "Paragraph as shown", "Paragraph solved"], start=1):
    cp.cell(row=1, column=i, value=h)
style_header(cp, 1, 5)
widths(cp, {"A": 6, "B": 26, "C": 10, "D": 78, "E": 78})
r = 2
for s in d["cloze"]:
    cp.cell(row=r, column=1, value=s["index"]).font = BASE
    cp.cell(row=r, column=2, value=s["title"]).font = BOLD
    cp.cell(row=r, column=3, value=f"{s['blanks'][0]['number']}–{s['blanks'][-1]['number']}").font = BASE
    cp.cell(row=r, column=4, value=s["prompt"]).font = BASE
    cp.cell(row=r, column=5, value=s["solved"]).font = BASE
    for c in range(1, 6):
        cell = cp.cell(row=r, column=c)
        cell.border = box
        cell.alignment = centre if c in (1, 3) else top_wrap
    cp.row_dimensions[r].height = 108
    r += 1

# ── Reading sections ───────────────────────────────────────────────────────
def reading_sheet(name, section):
    ws = wb.create_sheet(name)
    ws.sheet_view.showGridLines = False
    cols = ["#", "Set", "Passage", "Format", "Question type", "Question",
            "A", "B", "C", "D", "Key", "Correct answer", "Explanation"]
    for i, h in enumerate(cols, start=1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, 1, len(cols))
    widths(ws, {"A": 6, "B": 6, "C": 26, "D": 12, "E": 17, "F": 46,
                "G": 38, "H": 38, "I": 38, "J": 38, "K": 6, "L": 40, "M": 86})
    r = 2
    for s in [x for x in d["reading"] if x["section"] == section]:
        for q in s["questions"]:
            ws.cell(row=r, column=1, value=q["number"]).font = BASE
            ws.cell(row=r, column=2, value=s["index"]).font = BASE
            ws.cell(row=r, column=3, value=s["title"]).font = BASE
            ws.cell(row=r, column=4, value=s["format"]).font = BASE
            ws.cell(row=r, column=5, value=q["archetype"].replace("-", " ")).font = BASE
            ws.cell(row=r, column=6, value=q["prompt"]).font = BASE
            for off, cid in enumerate("abcd"):
                ws.cell(row=r, column=7 + off, value=q["choices"][cid]).font = BASE
            ws.cell(row=r, column=11, value=q["answer"].upper()).font = KEY
            # The answer text is looked up from the key, so the two can never disagree.
            ws.cell(row=r, column=12,
                    value=f"=INDEX(G{r}:J{r},MATCH(K{r},$G$1:$J$1,0))").font = Font(
                        name="Arial", size=10, color=GREEN)
            ws.cell(row=r, column=13, value=q["explanation"]).font = BASE
            for c in range(1, len(cols) + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = box
                cell.alignment = centre if c in (1, 2, 11) else top_wrap
                if s["index"] % 2 == 0:
                    cell.fill = band_fill
            ws.row_dimensions[r].height = 76
            r += 1
    ws.auto_filter.ref = f"A1:M{r-1}"
    return ws

reading_sheet("Read in Daily Life", "daily-life")
reading_sheet("Read an Academic Passage", "academic")

# ── Passages ───────────────────────────────────────────────────────────────
pg = wb.create_sheet("Passages")
pg.sheet_view.showGridLines = False
for i, h in enumerate(["Set", "Section", "Title", "Format", "Directions", "Questions", "Full text"], start=1):
    pg.cell(row=1, column=i, value=h)
style_header(pg, 1, 7)
widths(pg, {"A": 8, "B": 24, "C": 30, "D": 13, "E": 34, "F": 12, "G": 110})
label = {"daily-life": "Read in Daily Life", "academic": "Read an Academic Passage"}
r = 2
for s in d["reading"]:
    pg.cell(row=r, column=1, value=s["id"]).font = MONO
    pg.cell(row=r, column=2, value=label[s["section"]]).font = BASE
    pg.cell(row=r, column=3, value=s["title"]).font = BOLD
    pg.cell(row=r, column=4, value=s["format"]).font = BASE
    pg.cell(row=r, column=5, value=s["directions"]).font = BASE
    pg.cell(row=r, column=6,
            value=f"{s['questions'][0]['number']}–{s['questions'][-1]['number']}").font = BASE
    pg.cell(row=r, column=7, value=s["passage"]).font = BASE
    for c in range(1, 8):
        cell = pg.cell(row=r, column=c)
        cell.border = box
        cell.alignment = centre if c in (1, 4, 6) else top_wrap
    pg.row_dimensions[r].height = 150
    r += 1

wb.calculation.fullCalcOnLoad = True
wb.save(OUT)
print("written", OUT)
