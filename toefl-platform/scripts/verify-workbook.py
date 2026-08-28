"""Independent check of the workbook.

LibreOffice cannot run in this sandbox, so recalc.py is unavailable. Instead:
read every formula back, resolve its references against the cells it actually
points at, and confirm the result matches the source data. A wrong range shows
up here exactly as it would in Excel.
"""
import json
import re

from openpyxl import load_workbook

PATH = "/home/user/saas-landing/toefl-platform/docs/TOEFL-Reading-Question-Bank.xlsx"
bank = json.load(open("/tmp/claude-0/-home-user-saas-landing/cbf105b2-215c-546e-a953-9c154a810568/scratchpad/bank.json"))

wb = load_workbook(PATH)
problems = []
checked = 0

# ── expected truth, straight from the source data ──────────────────────────
exp_cloze = sum(len(s["blanks"]) for s in bank["cloze"])
exp_daily = sum(len(s["questions"]) for s in bank["reading"] if s["section"] == "daily-life")
exp_acad = sum(len(s["questions"]) for s in bank["reading"] if s["section"] == "academic")
exp_sets = {
    "Complete the Words": (len(bank["cloze"]), exp_cloze),
    "Read in Daily Life": (len([s for s in bank["reading"] if s["section"] == "daily-life"]), exp_daily),
    "Read an Academic Passage": (len([s for s in bank["reading"] if s["section"] == "academic"]), exp_acad),
}

print("sheets:", wb.sheetnames)

# ── Overview: resolve each SUMPRODUCT/COUNT against the sheet it names ─────
ov = wb["Overview"]
for row in range(6, 9):
    name = ov.cell(row=row, column=2).value
    sets_f = ov.cell(row=row, column=3).value
    q_f = ov.cell(row=row, column=4).value
    exp_set_n, exp_q_n = exp_sets[name]

    m = re.match(r"=SUMPRODUCT\(1/COUNTIF\('(.+?)'!B2:B(\d+),'(.+?)'!B2:B(\d+)\)\)$", sets_f)
    assert m, f"unexpected sets formula: {sets_f}"
    sheet_a, last_a, sheet_b, last_b = m.group(1), int(m.group(2)), m.group(3), int(m.group(4))
    if sheet_a != name or sheet_b != name:
        problems.append(f"{name}: sets formula points at {sheet_a}/{sheet_b}")
    vals = [wb[sheet_a].cell(row=r, column=2).value for r in range(2, last_a + 1)]
    if any(v is None for v in vals):
        problems.append(f"{name}: sets range B2:B{last_a} runs past the data")
    got = len(set(vals))
    if got != exp_set_n:
        problems.append(f"{name}: distinct sets resolve to {got}, expected {exp_set_n}")
    checked += 1

    m = re.match(r"=COUNT\('(.+?)'!A2:A(\d+)\)$", q_f)
    assert m, f"unexpected questions formula: {q_f}"
    sheet, last = m.group(1), int(m.group(2))
    nums = [wb[sheet].cell(row=r, column=1).value for r in range(2, last + 1)]
    got = len([n for n in nums if isinstance(n, int)])
    if got != exp_q_n:
        problems.append(f"{name}: COUNT resolves to {got}, expected {exp_q_n}")
    # Compare against max_row rather than reading A(last+1): openpyxl *creates*
    # a cell on access, which would extend the sheet and corrupt this check.
    if wb[sheet].max_row != last:
        problems.append(f"{name}: sheet ends at row {wb[sheet].max_row}, range stops at {last}")
    checked += 1

# ── Complete the Words: =D&E must rebuild the real word ───────────────────
cw = wb["Complete the Words"]
flat = [b for s in bank["cloze"] for b in s["blanks"]]
if cw.max_row - 1 != len(flat):
    problems.append(f"Complete the Words: {cw.max_row - 1} rows, expected {len(flat)}")
for i, b in enumerate(flat):
    r = i + 2
    if cw.cell(row=r, column=1).value != b["number"]:
        problems.append(f"row {r}: number {cw.cell(row=r, column=1).value} != {b['number']}")
    f = cw.cell(row=r, column=6).value
    if f != f"=D{r}&E{r}":
        problems.append(f"row {r}: word formula is {f}")
        continue
    got = str(cw.cell(row=r, column=4).value) + str(cw.cell(row=r, column=5).value)
    if got != b["word"]:
        problems.append(f"row {r}: =D&E resolves to '{got}', expected '{b['word']}'")
    checked += 1

# ── Reading sheets: INDEX/MATCH must return the correct choice text ───────
for sheet_name, section in [("Read in Daily Life", "daily-life"), ("Read an Academic Passage", "academic")]:
    ws = wb[sheet_name]
    headers = [ws.cell(row=1, column=c).value for c in range(7, 11)]
    if headers != ["A", "B", "C", "D"]:
        problems.append(f"{sheet_name}: choice headers are {headers}, MATCH will fail")
    qs = [q for s in bank["reading"] if s["section"] == section for q in s["questions"]]
    if ws.max_row - 1 != len(qs):
        problems.append(f"{sheet_name}: {ws.max_row - 1} rows, expected {len(qs)}")
    for i, q in enumerate(qs):
        r = i + 2
        if ws.cell(row=r, column=1).value != q["number"]:
            problems.append(f"{sheet_name} row {r}: number mismatch")
        f = ws.cell(row=r, column=12).value
        if f != f"=INDEX(G{r}:J{r},MATCH(K{r},$G$1:$J$1,0))":
            problems.append(f"{sheet_name} row {r}: lookup formula is {f}")
            continue
        key = ws.cell(row=r, column=11).value
        if key not in ("A", "B", "C", "D"):
            problems.append(f"{sheet_name} row {r}: key '{key}' is not A–D")
            continue
        got = ws.cell(row=r, column=7 + headers.index(key)).value
        if got != q["answerText"]:
            problems.append(f"{sheet_name} row {r}: lookup resolves to '{got}', expected '{q['answerText']}'")
        if key.lower() != q["answer"]:
            problems.append(f"{sheet_name} row {r}: key {key} != source answer {q['answer']}")
        checked += 1

# ── every formula in the book must be one of the three known shapes ───────
known = (re.compile(r"^=D\d+&E\d+$"),
         re.compile(r"^=INDEX\(G\d+:J\d+,MATCH\(K\d+,\$G\$1:\$J\$1,0\)\)$"),
         re.compile(r"^=SUMPRODUCT\(1/COUNTIF\(.+\)\)$"),
         re.compile(r"^=COUNT\(.+\)$"),
         re.compile(r"^=SUM\([A-Z]\d+:[A-Z]\d+\)$"))
total_formulas = 0
for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                total_formulas += 1
                if not any(k.match(cell.value) for k in known):
                    problems.append(f"{ws.title}!{cell.coordinate}: unrecognised formula {cell.value}")

# ── passages sheet completeness ──────────────────────────────────────────
pg = wb["Passages"]
if pg.max_row - 1 != len(bank["reading"]):
    problems.append(f"Passages: {pg.max_row - 1} rows, expected {len(bank['reading'])}")
for i, s in enumerate(bank["reading"]):
    if pg.cell(row=i + 2, column=7).value != s["passage"]:
        problems.append(f"Passages row {i+2}: text does not match {s['id']}")

cp = wb["Cloze Paragraphs"]
if cp.max_row - 1 != len(bank["cloze"]):
    problems.append(f"Cloze Paragraphs: {cp.max_row - 1} rows, expected {len(bank['cloze'])}")

print(f"formulas: {total_formulas} total, {checked} resolved against their data")
print(f"recalc on open: {wb.calculation.fullCalcOnLoad}")
print("PROBLEMS:\n" + "\n".join(problems) if problems else "✓ every formula resolves to the correct value")
